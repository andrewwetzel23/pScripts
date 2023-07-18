import openpyxl
import sys
from tqdm import tqdm

sys.path.append('..')
from pyFuncs import browseForFile

def update_translations(workbook):
    # Get the "Sheet" worksheet
    sheet = workbook['Sheet']

    # Extract English phrases and their Spanish translations from the "Sheet" worksheet
    translations = {}
    for row in sheet.iter_rows():
        english_phrase = row[0].value
        spanish_translation = row[1].value
        if english_phrase and spanish_translation:
            translations[english_phrase] = spanish_translation

    # Create a text file to store the missing translations
    missing_translations_file = open('missing_translations.txt', 'w')

    # Iterate over the other sheets in the workbook
    for worksheet_name in tqdm(workbook.sheetnames):
        if worksheet_name == 'Sheet':
            continue  # Skip the "Sheet" worksheet
        worksheet = workbook[worksheet_name]

        # Iterate over the rows and update with Spanish translations
        for row in worksheet.iter_rows(min_row=2):
            english_phrase = row[0].value
            if english_phrase:
                if english_phrase in translations:
                    spanish_translation = translations[english_phrase]
                    worksheet.cell(row=row[0].row, column=2).value = spanish_translation
                else:
                    missing_translations_file.write(f"{worksheet_name}: {english_phrase}\n")

    # Save the modified workbook
    output_file = 'output.xlsx'
    workbook.save(output_file)
    print(f"Translations updated successfully. Translated spreadsheet saved as '{output_file}'.")

    # Close the missing translations file
    missing_translations_file.close()

if __name__ == '__main__':
    # Load the workbook
    workbook = openpyxl.load_workbook(browseForFile())
    
    update_translations(workbook)
