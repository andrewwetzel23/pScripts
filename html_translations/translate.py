import openpyxl
from googletrans import Translator

from funcs import searchForFile

# Load the Excel spreadsheet
file = searchForFile()
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

# Create a translator object
translator = Translator()

# Iterate over the rows and translate the English text to Spanish
for row in sheet.iter_rows(values_only=True):
    english_text = row[0]
    if english_text:
        translation = translator.translate(english_text, src='en', dest='es')
        spanish_text = translation.text
        sheet.cell(row=row[0].row, column=row[0].column + 1).value = spanish_text

# Save the translated spreadsheet
workbook.save('output.xlsx')
