import openpyxl
from bs4 import BeautifulSoup
import sys
sys.path.append('..')
from pyFuncs import browseForFile, browseForDir, getFilesFromDir
import openpyxl
import webbrowser
import os
from tqdm import tqdm
import concurrent.futures
import re


import openpyxl
from bs4 import BeautifulSoup
import re

import openpyxl
from bs4 import BeautifulSoup
import re

def parseHtml(excel_file, html_file):
    # Load the HTML file
    with open(html_file, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Create a BeautifulSoup object to parse the HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Create a new workbook or load an existing one
    workbook = create_or_load_workbook(excel_file)

    # Create a new sheet in the workbook with the HTML file name
    sheet_name = generateSheetName(html_file)
    sheet = workbook.create_sheet(title=sheet_name)

    # Find all text elements in the HTML file from the specified tags and write them to the sheet
    row_index = 1
    illegal_char_pattern = re.compile(r'[\\/:*?|]')
    group_tags = ['title', 'button', 'option', 'p', 'label', 'tr', 'td', 'th', 'led-indicator']
    for tag in group_tags:
        sheet.cell(row=row_index, column=1).value = tag.upper()
        row_index += 1
        elements = soup.find_all(tag)
        for element in elements:
            write_text_elements(sheet, row_index, element, illegal_char_pattern)
            row_index += 1

        add_empty_row(sheet, row_index)

    # Save the workbook to the Excel file
    workbook.save(excel_file)


def create_or_load_workbook(excel_file):
    if not excel_file.endswith('.xlsx'):
        excel_file += '.xlsx'
    workbook = openpyxl.Workbook()
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        pass
    return workbook

def generateSheetName(html_file):
    return os.path.basename(html_file).split('.')[0]
    

def write_text_elements(sheet, row_index, element, illegal_char_pattern):
    text = element.string
    if not text:
        text = element.get_text()
    if text and text[0] != '#':
        sanitized_text = re.sub(illegal_char_pattern, '', text.strip())
        sheet.cell(row=row_index, column=1).value = sanitized_text


def add_empty_row(sheet, row_index):
    sheet.insert_rows(row_index)


def generateSheetName(html_file):
    sheet_name = os.path.basename(html_file).split('/')[-1].split('.')[0]
    return sheet_name

def filterFilesByWord(files, word):
    filtered_files = []
    for file in files:
        filename = os.path.basename(file)
        if word not in filename:
            filtered_files.append(file)
    return filtered_files

def consolidate_excel_sheets(excel_file):
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Get the first sheet
    first_sheet = workbook.worksheets[0]

    # Clear the existing content in the first sheet
    clear_sheet(first_sheet)

    # Iterate over the sheets (excluding the first sheet)
    for sheet in workbook.worksheets[1:]:
        # Extract unique cell values from the sheet
        unique_values = extract_unique_values(sheet)

        # Paste the unique cell values into the first column of the first sheet
        paste_values_to_first_sheet(first_sheet, unique_values)

    # Save the modified workbook
    workbook.save(excel_file)

def clear_sheet(sheet):
    # Clear all cells in the sheet except for the first column
    for row in sheet.iter_rows(min_row=2):
        for cell in row[1:]:
            cell.value = None

def extract_unique_values(sheet):
    # Extract unique cell values from the sheet's first column
    values = set()
    for cell in sheet.iter_cols(min_col=1, max_col=1, values_only=True):
        if cell[0]:
            values.add(cell[0])
    return list(values)

def paste_values_to_first_sheet(first_sheet, values):
    # Paste the values into the first column of the first sheet
    column = first_sheet['A']
    for i, value in enumerate(values, start=1):
        column[i].value = value


def main():
    excel_file = 'hmiTranslations.xlsx'
    if os.path.exists(excel_file):
        os.remove(excel_file)
    workbook = openpyxl.Workbook()
    workbook.save(excel_file)

    dir = browseForDir()
    html_files = getFilesFromDir(dir, '.html', True)
    no_spanish_html_files = filterFilesByWord(html_files, 'spanish')

    for file in tqdm(no_spanish_html_files):
        parseHtml(excel_file, file)

    consolidate_excel_sheets(excel_file)

if __name__ == "__main__":
    main()